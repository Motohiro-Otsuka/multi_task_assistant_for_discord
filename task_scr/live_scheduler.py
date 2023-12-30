import datetime
from zoneinfo import ZoneInfo
import discord
from discord.ext import commands
import tempfile

import os
import openai
import json
from PIL import Image, ImageFilter,  ImageDraw
import shutil
#from openpyxl import Workbook
import openpyxl


from pydrive2.auth import GoogleAuth
from pydrive2.drive import GoogleDrive

import time

class LiveScheduer:
    def __init__(self,config,use_drive,gdrive_setting_path=""):
        self.use_drive = use_drive
        if(self.use_drive):
            gauth = GoogleAuth(gdrive_setting_path) #if gdrive_setting_path != "" else GoogleAuth()
            gauth.LocalWebserverAuth()
            self.drive = GoogleDrive(gauth)
        else:
            pass
        self.config = config
        self.tmp_dir_path = self.config["files"]["tmp_dir_path"]
        self.icons = self.config["files"]["icons"]
        self.contents=self.config["files"]["contents"]
        self.platform = self.config["files"]["platform"]
        self.schedule = self.config["files"]["schedule"]
        self.number_img = self.config["files"]["number_img"]
        self.column_size = self.config["grid"]["column_size"]
        self.column_first_point = self.config["grid"]["column_first_point"]
        self.editing_dic = {}#{"thread-id":{"schedule":{},status=["choice","edit","end"],day="",timestamp:""}}

    def check_schedule_editing(self,thread_id):
        """
        すでにchatのスレッドが立っているか確認するための
        """
        if (thread_id in self.editing_dic):
            return True
        else:
            return False


    def centering(self,img,column_key):
        x,y = img.size
        column_x = self.column_size[column_key]["x"]
        column_y = self.column_size[column_key]["y"]
        return int ((column_x-x)/2),int ((column_y-y)/2)

    async def return_schedule_picture(self,img,ctx):
        if (ctx == None):
            img.show()
        else:
            os.makedirs(self.tmp_dir_path, exist_ok=True)
            img_path = os.path.join(self.tmp_dir_path,"schedule.png")
            img.save(img_path)
            await ctx.send(file=discord.File(img_path))

    async def print_schedule(self,ctx=None):
        schedule_obj = self.get_schedule()
        base_image = self.get_schedule_base_image()
        liver_images = self.get_liver_icon()
        platform_images = self.get_live_platform_icon()
        contents_images = self.get_live_contents_icon()
        number_picture = self.get_number_picture()
        #draw = ImageDraw.Draw(base_image)
        i = 0
        for key,val in schedule_obj.items():
            for column_key,column_point in self.column_first_point.items():
                loop = 1 if column_key != "liver" else len(val[column_key].keys())
                size = self.column_size[column_key]
                if (column_key == "day"):
                    j = 0
                    for d in key:
                        number_img = self.image_resize(number_picture[d],column_key)
                        c_x,c_y = self.centering(number_img,column_key)
                        if(j==1):
                            c_x = int(c_x * -1)
                        base_image.paste(number_img,
                                        (column_point["x"]+size["x"]*(j)+c_x,column_point["y"]+size["y"]*i+c_y),
                                        number_img)
                        j += 1    
                elif (column_key == "time" and val["開始時間"] is not None):
                    start_time = val["開始時間"].strftime("%H:%M")
                    j = 0
                    for d in start_time:
                        number_img = self.image_resize(number_picture[d],column_key)
                        c_x,c_y = self.centering(number_img,column_key)
                        if(j==1):
                            c_x = int(c_x * -1)
                        base_image.paste(number_img,
                                        (column_point["x"]+size["x"]*(j)+c_x,column_point["y"]+size["y"]*i+c_y),
                                        number_img)
                        j += 1
                elif(column_key == "content" and val["内容"] is not None):
                    c_x,c_y = self.centering(contents_images[val["内容"]],column_key)
                    base_image.paste(contents_images[val["内容"]] ,
                                    (column_point["x"]+c_x,column_point["y"]+size["y"]*i+c_y),
                                    contents_images[val["内容"]])
                elif(column_key == "liver"):
                    j = 0
                    for liver_name,state in val["liver"].items():
                        if(state == 'o'):
                            c_x,c_y = self.centering(liver_images[liver_name],column_key)
                            base_image.paste(liver_images[liver_name],
                                            (column_point["x"]+size["x"]*(j)+c_x,column_point["y"]+size["y"]*i+c_y),
                                            liver_images[liver_name])
                        j += 1
                elif(column_key == "platform" and val["サイト"] is not None):
                    c_x,c_y = self.centering(platform_images[val["サイト"]],column_key)
                    base_image.paste(platform_images[val["サイト"]] ,
                                    (column_point["x"]+c_x,column_point["y"]+size["y"]*i+c_y))
            i += 1

        await self.return_schedule_picture(base_image,ctx)

    async def print_grid_schedule_baseimg(self,ctx=None):
        base_image = self.get_schedule_base_image()
        img_width,img_height = base_image.size
        draw = ImageDraw.Draw(base_image)
        for i in range(0,img_width,10):#10pxごとに縦線を引く
            if(i % 50 == 0):
                draw.line((i,0 ,i ,img_height), fill=(255, 0, 0), width=1)
            else:
                draw.line((i,0 ,i ,img_height), fill=(240, 230, 140), width=1)
        for i in range(0,img_height,10):#10pxごとに横線を引く
            if(i % 50 == 0):
                draw.line((0,i ,img_width ,i), fill=(255, 0, 0), width=1)
            else:
                draw.line((0,i ,img_width ,i), fill=(240, 230, 140), width=1)
        await self.return_schedule_picture(base_image,ctx)

    def image_resize(self,img,column_name):
        img_x,img_y = img.size
        column_x_size = self.column_size[column_name]["x"]
        column_y_size = self.column_size[column_name]["y"]
        content_scale = column_x_size / img_x if  column_x_size / img_x < column_y_size / img_y else column_y_size / img_y
        content_scale = content_scale-0.05
        return img.resize((int(img_x*content_scale),int(img_y*content_scale)))

    def get_image(self,img_id):
        if(self.use_drive):
            img = self.drive.CreateFile({'id': img_id})
            os.makedirs(self.tmp_dir_path, exist_ok=True)
            img_path = os.path.join(self.tmp_dir_path,"{}.png".format(img_id))
            img.GetContentFile(img_path)
        else:
            img_path = img_id
        image_obj = Image.open(img_path)
        return image_obj

    def get_schedule_base_image(self):
        return self.get_image(self.config["files"]["format"].split("/")[5])
    
    def get_icons(self,icon_paths,grid_name):
        icon_objects = {}
        for key,val in icon_paths.items():
            if(grid_name == "number_img"):
                if(self.use_drive):
                    icon_objects[key]=self.get_image(val.split("/")[5])
                else:
                    icon_objects[key]=self.get_image(val)
            else:
                if(self.use_drive):
                    icon_objects[key]=self.image_resize(self.get_image(val.split("/")[5]),grid_name)
                else:
                    icon_objects[key]=self.image_resize(self.get_image(val),grid_name)
        return icon_objects

    def get_liver_icon(self):
        return self.get_icons(self.icons,"liver")
        
    def get_live_contents_icon(self):
        return self.get_icons(self.contents,"content")

    def get_live_platform_icon(self):
        return self.get_icons(self.platform,"platform")

    def get_number_picture(self):
        return self.get_icons(self.number_img,"number_img")

    def schedule_json_to_markdown(self,date,date_item):
        response = "- {}日\n".format(date)
        for column,val in date_item.items():
            if(column=="liver"):
                response += "  - liver\n" 
                for name,assign in val.items():
                    if(assign is None):
                        response += "    - {}:{}\n".format(name,"x")
                    else:
                        response += "    - {}:{}\n".format(name,assign)
            elif(column=="開始時間" or column=="終了時間"):
                if(val is not None):
                    response += "  - {}:{}\n".format(column,val.strftime("%H:%M")) 
                else:
                    response += "  - {}:{}\n".format(column,"")
            else:
                if(val is not None):
                    response += "  - {}:{}\n".format(column,val)
                else:
                    response += "  - {}:{}\n".format(column,"")
        return response
        
    def schedule_all_json_to_markdown(self,schedule=None):
        schedule_obj = self.get_schedule() if schedule==None else schedule
        response = ""
        for date,date_item in schedule_obj.items():
            response += self.schedule_json_to_markdown(date,date_item)
        return response

    def schedule_markdown_to_json(self,text):
        result = {}
        liver={}
        for line in text.split("\n")[1:]:
            line_list = line.split(" ")
            if(len(line_list)==2):#日付
                pass
            elif(len(line_list)==4):
                if(line_list[-1]=="liver"):
                    result[line_list[-1]] = {}
                else:
                    tmp = line_list[-1].split(":")
                    if(len(tmp)==3):
                        result[tmp[0]]=datetime.datetime.strptime("{}:{}".format(tmp[1],tmp[2]),"%H:%M")
                    else:
                        result[tmp[0]]=tmp[1]
            elif(len(line_list)==6):
                tmp = line_list[-1].split(":")
                liver[tmp[0]]=tmp[1]
        result["liver"] = liver
        return result 

    async def show_schedule_data(self,ctx=None):
        response = self.schedule_all_json_to_markdown()
        channel = ctx.channel #メッセージがあったチャンネルの取得
        #スレッド名の定義
        date = datetime.datetime.now(ZoneInfo("Asia/Tokyo"))
        date_str = date.strftime('%Y-%m-%d %H:%M')
        thread_name = '{}現在のスケジュール'.format(date_str)
        #スレッドの作成
        thread = await channel.create_thread(name=thread_name,reason="スケジュール確認",type=discord.ChannelType.public_thread,auto_archive_duration=60)#スレッドを作る
        # メッセージを送信してスレッドを開始します。
        await thread.send(response)

    def get_schedule(self):#エクセルファイルからスケジュールを取得する
        def load_excel(excel_file_path):
            wb = openpyxl.load_workbook(excel_file_path)
            ws = wb.worksheets[0]
            #labelの取得
            for i in range(4):
                content = ws.cell(2, label_start_idx+i)#C2セルから
                label.append(content.value)
            #ライバーの取得
            i = 0
            while True:
                liver_name = ws.cell(2, liver_start_idx+i)#G2セルから
                if(liver_name.value == None):
                    break
                liver.append(liver_name.value)
                i += 1
            #スケジュールの取得
            for i in range(7):
                date = str(int(ws.cell(3+i, 2).value))
                schedule[date] = {"liver":{}}
                #配信情報を取得
                for j in range(len(label)):
                    content = ws.cell(3+i, label_start_idx+j)
                    schedule[date][label[j]] = content.value
                for j in range(len(liver)):
                    liver_state = ws.cell(3+i, liver_start_idx+j)
                    schedule[date]["liver"][liver[j]] = liver_state.value
            wb.close()
            return schedule
        schedule = {}
        label = []
        liver = []
        label_start_idx = 3
        liver_start_idx = 7
        if(self.use_drive):
            excel_file_id = self.schedule.split("/")[5]
            excle_wb = self.drive.CreateFile({'id': excel_file_id})
            with tempfile.TemporaryDirectory() as temp_dir:
                excel_file_path = '{}/schedule.xlsx'.format(temp_dir)
                excle_wb.GetContentFile(excel_file_path)
                return load_excel(excel_file_path)
        else:
            excel_file_path = self.schedule
            return load_excel(excel_file_path)

    async def save_new_schedule(self,trehad_id,new_sheet=False):
        def edit_excel_schedule(excel_file_path):
            label_start_idx = 3
            liver_start_idx = 7
            wb = openpyxl.load_workbook(excel_file_path)
            ws = wb.worksheets[0]
            schedule_items = self.editing_dic[trehad_id]["schedule"].items()
            i = 0
            for day, items in schedule_items:
                for key, val in items.items():
                    if(key == "開始時間"):#C列 = 3
                        ws.cell(i+3, 3).value = val
                    elif(key == "終了時間"):#D列 = 4
                        ws.cell(i+3, 4).value = val
                    elif(key == "内容"):#E列 = 5
                        ws.cell(i+3, 5).value = val
                    elif(key == "サイト"):#F列 = 6
                        ws.cell(i+3, 6).value = val
                    elif(key == "liver"):#G列以降 = 7以降
                        j = 0
                        for liver,assign in val.items():
                            ws.cell(i+3, 7+j).value = assign
                            j+=1
                i += 1
            wb.save(excel_file_path)
            wb.close()
        if(self.use_drive):
            excel_file_id = self.schedule.split("/")[5]
            excle_wb = self.drive.CreateFile({'id': excel_file_id})
            os.makedirs(self.tmp_dir_path, exist_ok=True)
            excel_file_path = os.path.join(self.tmp_dir_path,"schedule.xlsx")
            excle_wb.GetContentFile(excel_file_path)
            edit_excel_schedule(excel_file_path)
            excle_wb.SetContentFile(excel_file_path)
            excle_wb.Upload()
        else:
            excel_file_path = self.schedule
            edit_excel_schedule(excel_file_path)

    async def edit_schedule(self,thread,text=""):
        if(thread.id in self.editing_dic):#すでに編集中の時
            if(self.editing_dic[thread.id]["status"] == "choice"):#日付選択直後
                if(text in self.editing_dic[thread.id]["schedule"]):
                    choiced_schedule = self.schedule_json_to_markdown(int(text),self.editing_dic[thread.id]["schedule"][text])
                    self.editing_dic[thread.id]["status"] = "edit"
                    self.editing_dic[thread.id]["editing_day"]=text
                    await thread.send("次の箇条書きをコピーして、変更しスレッドに送信してください")
                    await thread.send(choiced_schedule)
                else:
                    await thread.send("{}日は対象外です。\n日付を選びなおしてください。".format(text))
            elif(self.editing_dic[thread.id]["status"] == "edit"):#日付選択直後
                new_schedule = self.schedule_markdown_to_json(text)
                day = self.editing_dic[thread.id]["editing_day"] 
                self.editing_dic[thread.id]["schedule"][day] = new_schedule
                await thread.send(self.schedule_all_json_to_markdown(self.editing_dic[thread.id]["schedule"]))
                await thread.send("変更後は上記のようになりました。")
                await self.save_new_schedule(thread.id)
                await thread.send("別の日程を変更する場合はもう一度schedule-editコマンドを叩いてください")
                del self.editing_dic[thread.id]
            #入力されたmarkdownを処理してスケジュールを更新する
            #先にスケジュールを作るコードを実装する
        else:#初めての場合は1週間のスケジュールを返す
            schedule_obj = self.get_schedule()
            self.editing_dic[thread.id] = {"schedule":schedule_obj,"status":"choice","editing_day":None,"timestamp":datetime.datetime.now()}
            await thread.send(self.schedule_all_json_to_markdown())
            await thread.send("変更したい日の日付を数字で答えてください")

if __name__ == "__main__":#検証用コード
    config_file = open("../config/config.json","r",encoding="utf-8")
    config = json.load(config_file)
    function_config = config["function"]
    live_scheduler = LiveScheduer(function_config["live_scheduler"])
    live_scheduler.show_schedule_data()
    shutil.rmtree("tmp")